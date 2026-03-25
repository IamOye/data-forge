"""
harvest_analytics.py — YouTube Analytics Data Harvester

Pulls channel and per-video stats from the YouTube Data API v3 and
YouTube Analytics API v2, then saves results to CSV and/or Excel.

Usage:
    .venv/Scripts/python.exe scripts/harvest_analytics.py
    .venv/Scripts/python.exe scripts/harvest_analytics.py --channel money_debate
    .venv/Scripts/python.exe scripts/harvest_analytics.py --format csv

Scheduler integration (src/scheduler.py run_daily_analytics):
    from scripts.harvest_analytics import harvest
    harvest(channel="money_debate")
"""

import argparse
import json
import logging
import re
import sys
from dataclasses import dataclass, field
from datetime import date, timedelta
from pathlib import Path
from typing import Any

from dotenv import load_dotenv

load_dotenv()

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

CREDENTIALS_DIR = Path(".credentials")
DEFAULT_OUTPUT_DIR = Path("data/analytics")
SHORTS_MAX_SECONDS = 60  # videos <= this are classified as "Short"

# CSV column names — order defines column order in output files
CHANNEL_CSV_COLUMNS = [
    "date", "subscribers", "total_views", "total_videos", "watch_time_hours",
]
VIDEO_CSV_COLUMNS = [
    "video_id", "title", "published_at", "duration_seconds", "type",
    "views", "likes", "comments", "favourites",
    "impression_count", "click_through_rate",
    "average_view_duration", "average_view_percentage",
    "shares", "subscribers_gained", "revenue",
    "tags", "description_length", "category_id",
]


# ---------------------------------------------------------------------------
# Data models
# ---------------------------------------------------------------------------

@dataclass
class ChannelStats:
    """Snapshot of channel-level statistics."""

    date: str
    subscribers: int = 0
    total_views: int = 0
    total_videos: int = 0
    watch_time_hours: float = 0.0
    channel_id: str = ""
    channel_title: str = ""

    def to_csv_row(self) -> dict[str, Any]:
        return {
            "date":             self.date,
            "subscribers":      self.subscribers,
            "total_views":      self.total_views,
            "total_videos":     self.total_videos,
            "watch_time_hours": self.watch_time_hours,
        }


@dataclass
class VideoRow:
    """Per-video metrics row."""

    video_id: str
    title: str = ""
    published_at: str = ""
    duration_seconds: int = 0
    type: str = ""                     # "Short" or "Long-form"
    views: int = 0
    likes: int = 0
    comments: int = 0
    favourites: int = 0
    impression_count: int = 0
    click_through_rate: float = 0.0
    average_view_duration: float = 0.0
    average_view_percentage: float = 0.0
    shares: int = 0
    subscribers_gained: int = 0
    revenue: str = ""                  # empty string when not monetized
    tags: str = ""                     # comma-joined tag list
    description_length: int = 0
    category_id: str = ""

    @property
    def like_rate(self) -> float:
        """likes / views, or 0.0 when views == 0."""
        return self.likes / self.views if self.views > 0 else 0.0

    def to_dict(self) -> dict[str, Any]:
        return {
            "video_id":               self.video_id,
            "title":                  self.title,
            "published_at":           self.published_at,
            "duration_seconds":       self.duration_seconds,
            "type":                   self.type,
            "views":                  self.views,
            "likes":                  self.likes,
            "comments":               self.comments,
            "favourites":             self.favourites,
            "impression_count":       self.impression_count,
            "click_through_rate":     self.click_through_rate,
            "average_view_duration":  self.average_view_duration,
            "average_view_percentage": self.average_view_percentage,
            "shares":                 self.shares,
            "subscribers_gained":     self.subscribers_gained,
            "revenue":                self.revenue,
            "tags":                   self.tags,
            "description_length":     self.description_length,
            "category_id":            self.category_id,
        }


@dataclass
class HarvestResult:
    """Result of a single harvest run."""

    channel_key: str
    videos_count: int = 0
    total_views: int = 0
    subscribers: int = 0
    output_dir: Path = field(default_factory=lambda: DEFAULT_OUTPUT_DIR)
    saved_files: list[str] = field(default_factory=list)
    is_valid: bool = True
    errors: list[str] = field(default_factory=list)

    def summary(self) -> str:
        return (
            f"Harvested: {self.videos_count} videos, "
            f"{self.total_views:,} total views, "
            f"{self.subscribers:,} subscribers, "
            f"saved to {self.output_dir}"
        )

    def to_dict(self) -> dict[str, Any]:
        return {
            "channel_key":  self.channel_key,
            "videos_count": self.videos_count,
            "total_views":  self.total_views,
            "subscribers":  self.subscribers,
            "output_dir":   str(self.output_dir),
            "saved_files":  self.saved_files,
            "is_valid":     self.is_valid,
            "errors":       self.errors,
        }


# ---------------------------------------------------------------------------
# AnalyticsHarvester
# ---------------------------------------------------------------------------

class AnalyticsHarvester:
    """
    Fetches YouTube channel and video analytics then writes CSV / Excel output.

    Google API and openpyxl are lazy-imported so the module loads without
    those packages installed.

    Args:
        channel_key:     Selects which OAuth token to load from credentials_dir.
        credentials_dir: Directory containing {channel_key}_token.json files.
        output_dir:      Directory where output files are written.
    """

    def __init__(
        self,
        channel_key: str = "money_debate",
        credentials_dir: str | Path = CREDENTIALS_DIR,
        output_dir: str | Path = DEFAULT_OUTPUT_DIR,
    ) -> None:
        self.channel_key = channel_key
        self.credentials_dir = Path(credentials_dir)
        self.output_dir = Path(output_dir)

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def harvest(self, format: str = "both") -> HarvestResult:
        """
        Run the full harvest pipeline.

        1. Fetch channel-level stats (channels.list).
        2. Fetch all uploaded video metadata (playlistItems + videos.list).
        3. Enrich each video with analytics (YouTube Analytics API v2).
        4. Save requested output formats (CSV and/or Excel).

        Args:
            format: "csv", "excel", or "both".

        Returns:
            HarvestResult. Never raises — errors are captured inside.
        """
        today = date.today().isoformat()
        result = HarvestResult(channel_key=self.channel_key, output_dir=self.output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # ── Channel stats ──────────────────────────────────────────────
        channel_stats = ChannelStats(date=today)
        try:
            channel_stats = self._fetch_channel_stats(today)
        except Exception as exc:
            logger.warning("[harvest] Could not fetch channel stats: %s", exc)
            result.errors.append(f"channel_stats: {exc}")

        result.subscribers = channel_stats.subscribers
        result.total_views = channel_stats.total_views

        # ── Video list ─────────────────────────────────────────────────
        videos: list[VideoRow] = []
        try:
            videos = self._fetch_all_videos()
        except Exception as exc:
            logger.warning("[harvest] Could not fetch video list: %s", exc)
            result.errors.append(f"videos: {exc}")

        # ── Analytics enrichment (best-effort per video) ───────────────
        for video in videos:
            try:
                self._enrich_video_analytics(video)
            except Exception as exc:
                logger.warning(
                    "[harvest] Analytics enrichment failed for %s: %s",
                    video.video_id, exc,
                )

        result.videos_count = len(videos)

        # ── Save outputs ───────────────────────────────────────────────
        if format in ("csv", "both"):
            try:
                ch_file = self._save_channel_csv(channel_stats, today)
                vid_file = self._save_video_csv(videos, today)
                result.saved_files.extend([ch_file, vid_file])
            except Exception as exc:
                logger.error("[harvest] CSV save failed: %s", exc)
                result.errors.append(f"csv: {exc}")

        if format in ("excel", "both"):
            try:
                xl_file = self._save_excel(channel_stats, videos, today)
                result.saved_files.append(xl_file)
            except Exception as exc:
                logger.error("[harvest] Excel save failed: %s", exc)
                result.errors.append(f"excel: {exc}")

        # ── GSheet sync (best-effort) ──────────────────────────────────
        if videos:
            try:
                self.sync_to_gsheet(videos)
            except Exception as exc:
                logger.warning("[harvest] GSheet sync failed: %s", exc)
                result.errors.append(f"gsheet: {exc}")

        result.is_valid = len(result.errors) == 0
        logger.info("[harvest] %s", result.summary())
        return result

    # ------------------------------------------------------------------
    # Pure static helpers — fully testable without mocking
    # ------------------------------------------------------------------

    @staticmethod
    def parse_iso8601_duration(duration: str) -> int:
        """
        Convert an ISO 8601 duration string to total seconds.

        Examples:
            "PT15S"     → 15
            "PT1M"      → 60
            "PT1M30S"   → 90
            "PT1H"      → 3600
            "PT2H3M4S"  → 7384
            "P1D"       → 86400
            ""          → 0
            "P0D"       → 0
        """
        if not duration:
            return 0
        m = re.match(
            r"P(?:(\d+)D)?(?:T(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?)?",
            duration,
        )
        if not m:
            return 0
        days    = int(m.group(1) or 0)
        hours   = int(m.group(2) or 0)
        minutes = int(m.group(3) or 0)
        seconds = int(m.group(4) or 0)
        return days * 86_400 + hours * 3_600 + minutes * 60 + seconds

    @staticmethod
    def detect_video_type(duration_seconds: int) -> str:
        """Return 'Short' if duration_seconds <= 60, else 'Long-form'."""
        return "Short" if duration_seconds <= SHORTS_MAX_SECONDS else "Long-form"

    # ------------------------------------------------------------------
    # Credential / service helpers (lazy Google imports)
    # ------------------------------------------------------------------

    def _credentials_path(self) -> Path:
        """Return the OAuth token path for the current channel_key."""
        return self.credentials_dir / f"{self.channel_key}_token.json"

    def _load_credentials(self):
        """
        Load OAuth2 credentials from .credentials/{channel_key}_token.json.

        Falls back to default_token.json if the channel-specific file is absent.
        Raises FileNotFoundError when neither file exists.
        """
        from google.oauth2.credentials import Credentials  # lazy

        token_path = self._credentials_path()
        if not token_path.exists():
            fallback = self.credentials_dir / "default_token.json"
            if fallback.exists():
                logger.debug(
                    "[harvest] %s not found — falling back to %s",
                    token_path.name, fallback.name,
                )
                token_path = fallback
            else:
                raise FileNotFoundError(
                    f"Credentials not found: {token_path} (also tried {fallback})"
                )

        data = json.loads(token_path.read_text())
        return Credentials(
            token=data.get("token"),
            refresh_token=data.get("refresh_token"),
            token_uri=data.get("token_uri", "https://oauth2.googleapis.com/token"),
            client_id=data.get("client_id"),
            client_secret=data.get("client_secret"),
            scopes=data.get("scopes"),
        )

    def _build_data_service(self, credentials):
        """Build YouTube Data API v3 service."""
        from googleapiclient.discovery import build  # lazy
        return build("youtube", "v3", credentials=credentials)

    def _build_analytics_service(self, credentials):
        """Build YouTube Analytics API v2 service."""
        from googleapiclient.discovery import build  # lazy
        return build("youtubeAnalytics", "v2", credentials=credentials)

    # ------------------------------------------------------------------
    # API fetchers
    # ------------------------------------------------------------------

    def _fetch_channel_stats(self, today: str) -> ChannelStats:
        """Fetch channel-level statistics via channels.list."""
        credentials = self._load_credentials()
        service = self._build_data_service(credentials)

        resp = service.channels().list(
            part="snippet,statistics",
            mine=True,
        ).execute()

        items = resp.get("items", [])
        if not items:
            logger.warning("[harvest] channels.list returned no items")
            return ChannelStats(date=today)

        item    = items[0]
        stats   = item.get("statistics", {})
        snippet = item.get("snippet", {})

        return ChannelStats(
            date=today,
            subscribers=int(stats.get("subscriberCount", 0)),
            total_views=int(stats.get("viewCount", 0)),
            total_videos=int(stats.get("videoCount", 0)),
            channel_id=item.get("id", ""),
            channel_title=snippet.get("title", ""),
        )

    def _fetch_all_videos(self) -> list[VideoRow]:
        """
        Fetch all uploaded videos via playlistItems.list → videos.list.

        Paginates automatically; batches video detail requests in groups of 50.
        """
        credentials = self._load_credentials()
        service = self._build_data_service(credentials)

        # Get the uploads playlist ID
        ch_resp = service.channels().list(
            part="contentDetails",
            mine=True,
        ).execute()

        items = ch_resp.get("items", [])
        if not items:
            return []

        uploads_playlist = (
            items[0]
            .get("contentDetails", {})
            .get("relatedPlaylists", {})
            .get("uploads", "")
        )
        if not uploads_playlist:
            logger.warning("[harvest] No uploads playlist found")
            return []

        # Collect all video IDs from the playlist (paginated)
        video_ids: list[str] = []
        page_token = None
        while True:
            kwargs: dict[str, Any] = {
                "part":       "contentDetails",
                "playlistId": uploads_playlist,
                "maxResults": 50,
            }
            if page_token:
                kwargs["pageToken"] = page_token

            pl_resp = service.playlistItems().list(**kwargs).execute()
            for item in pl_resp.get("items", []):
                vid_id = item.get("contentDetails", {}).get("videoId", "")
                if vid_id:
                    video_ids.append(vid_id)

            page_token = pl_resp.get("nextPageToken")
            if not page_token:
                break

        # Fetch video details in batches of 50
        rows: list[VideoRow] = []
        for i in range(0, len(video_ids), 50):
            batch = video_ids[i : i + 50]
            vid_resp = service.videos().list(
                part="snippet,statistics,contentDetails",
                id=",".join(batch),
            ).execute()
            for item in vid_resp.get("items", []):
                rows.append(self._parse_video_item(item))

        logger.info("[harvest] Fetched %d video(s)", len(rows))
        return rows

    def _parse_video_item(self, item: dict[str, Any]) -> VideoRow:
        """Build a VideoRow from a videos.list response item."""
        snippet = item.get("snippet", {})
        stats   = item.get("statistics", {})
        content = item.get("contentDetails", {})

        duration_str  = content.get("duration", "")
        duration_secs = self.parse_iso8601_duration(duration_str)
        tags          = snippet.get("tags") or []
        description   = snippet.get("description") or ""

        return VideoRow(
            video_id=item.get("id", ""),
            title=snippet.get("title", ""),
            published_at=snippet.get("publishedAt", ""),
            duration_seconds=duration_secs,
            type=self.detect_video_type(duration_secs),
            views=int(stats.get("viewCount", 0)),
            likes=int(stats.get("likeCount", 0)),
            comments=int(stats.get("commentCount", 0)),
            favourites=int(stats.get("favoriteCount", 0)),
            tags=",".join(tags),
            description_length=len(description),
            category_id=snippet.get("categoryId", ""),
        )

    def _enrich_video_analytics(self, video: VideoRow) -> None:
        """
        Fetch deeper analytics (avg view %, shares, etc.) and update the
        VideoRow in place.  Silently skips if API call fails.

        Note: impressions / impressionClickThroughRate are NOT available via
        the Analytics API v2 (YouTube Studio only) and are intentionally omitted.
        """
        credentials = self._load_credentials()
        service = self._build_analytics_service(credentials)

        end_date = date.today()
        # Use the video's actual publish date so we get full lifetime stats.
        # published_at is ISO 8601: "2024-03-15T10:30:00Z" — take first 10 chars.
        start_date_str = (video.published_at or "")[:10]
        if not start_date_str:
            start_date_str = (end_date - timedelta(days=365)).isoformat()

        resp = service.reports().query(
            ids="channel==MINE",
            startDate=start_date_str,
            endDate=end_date.isoformat(),
            metrics=(
                "averageViewDuration,averageViewPercentage,"
                "shares,subscribersGained,"
                "likes,comments,estimatedMinutesWatched"
            ),
            filters=f"video=={video.video_id}",
        ).execute()

        rows = resp.get("rows", [])
        row  = rows[0] if rows else []

        def _int(idx: int) -> int:
            return int(row[idx]) if len(row) > idx and row[idx] is not None else 0

        def _float(idx: int) -> float:
            return float(row[idx]) if len(row) > idx and row[idx] is not None else 0.0

        # Index order matches the metrics string above:
        # 0: averageViewDuration, 1: averageViewPercentage,
        # 2: shares, 3: subscribersGained, 4: likes, 5: comments,
        # 6: estimatedMinutesWatched (fetched but not stored separately)
        video.average_view_duration   = _float(0)
        video.average_view_percentage = _float(1)
        video.shares                  = _int(2)
        video.subscribers_gained      = _int(3)
        # likes/comments from Analytics API override the Data API values when present
        analytics_likes    = _int(4)
        analytics_comments = _int(5)
        if analytics_likes > 0:
            video.likes = analytics_likes
        if analytics_comments > 0:
            video.comments = analytics_comments

    # ------------------------------------------------------------------
    # CSV writers
    # ------------------------------------------------------------------

    def _save_channel_csv(self, stats: ChannelStats, today: str) -> str:
        """Append one row to channel_stats_{today}.csv."""
        import csv

        path = self.output_dir / f"channel_stats_{today}.csv"
        write_header = not path.exists()
        with open(path, "a", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=CHANNEL_CSV_COLUMNS)
            if write_header:
                writer.writeheader()
            writer.writerow(stats.to_csv_row())

        logger.info("[harvest] Saved channel stats: %s", path)
        return str(path)

    def _save_video_csv(self, videos: list[VideoRow], today: str) -> str:
        """Write all videos to video_metrics_{today}.csv (overwrites)."""
        import csv

        path = self.output_dir / f"video_metrics_{today}.csv"
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=VIDEO_CSV_COLUMNS)
            writer.writeheader()
            for v in videos:
                writer.writerow(v.to_dict())

        logger.info("[harvest] Saved video metrics: %s (%d rows)", path, len(videos))
        return str(path)

    # ------------------------------------------------------------------
    # Excel writer
    # ------------------------------------------------------------------

    def _save_excel(
        self,
        channel_stats: ChannelStats,
        videos: list[VideoRow],
        today: str,
    ) -> str:
        """Write MoneyHeresy_Analytics_{today}.xlsx with 3 sheets."""
        import openpyxl  # lazy
        from openpyxl.styles import Font, PatternFill
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.utils import get_column_letter

        channel_name = self._get_channel_name()
        path = self.output_dir / f"{channel_name}_Analytics_{today}.xlsx"

        wb = openpyxl.Workbook()

        # Sheet 1 — Channel Overview
        ws_ch = wb.active
        ws_ch.title = "Channel Overview"
        self._write_channel_overview_sheet(ws_ch, channel_stats)

        # Sheet 2 — Video Metrics
        ws_vid = wb.create_sheet("Video Metrics")
        self._write_video_metrics_sheet(ws_vid, videos)

        # Sheet 3 — Top Performers
        ws_top = wb.create_sheet("Top Performers")
        self._write_top_performers_sheet(ws_top, videos)

        wb.save(path)
        logger.info("[harvest] Saved Excel: %s", path)
        return str(path)

    def sync_to_gsheet(self, videos: list[VideoRow]) -> None:
        """Write analytics data to the 'Performance' tab in the ChannelForge GSheet.

        Tab columns (in order):
        video_id | Published | GSheet SEQ | Topic Brief | YouTube Title |
        Views | Likes | Watch% | Avg Duration | Subs Gained | Shares | Category

        Joins on youtube_video_id against production_results in channel_forge.db
        to get keyword (topic brief), and LEFT JOINs manual_topics for seq.
        YouTube Title comes from the live YouTube API data (VideoRow.title).
        Creates the tab if it does not exist.
        Clears and rewrites the full tab on each call.
        """
        import sqlite3 as _sq
        from src.crawler.gsheet_topic_sync import get_gsheet_client

        _, spreadsheet = get_gsheet_client()

        # Try to get or create Performance tab
        try:
            import gspread
            try:
                ws = spreadsheet.worksheet("Performance")
            except gspread.WorksheetNotFound:
                ws = spreadsheet.add_worksheet(title="Performance", rows=600, cols=12)
        except Exception as exc:
            logger.error("[harvest] Could not open/create Performance tab: %s", exc)
            return

        ws.clear()

        headers = [
            "video_id", "Published", "GSheet SEQ", "Topic Brief", "YouTube Title",
            "Views", "Likes", "Watch%", "Avg Duration", "Subs Gained", "Shares", "Category",
        ]
        ws.append_row(headers, value_input_option="USER_ENTERED")

        # Build video_id → (seq, keyword) lookup from production_results + manual_topics
        db_path = Path("data/processed/channel_forge.db")
        vid_lookup: dict[str, tuple[str, str]] = {}
        if db_path.exists():
            try:
                conn = _sq.connect(db_path)
                try:
                    # Check if manual_topics table exists for SEQ lookup
                    has_manual = conn.execute(
                        "SELECT name FROM sqlite_master "
                        "WHERE type='table' AND name='manual_topics'"
                    ).fetchone() is not None

                    if has_manual:
                        rows = conn.execute(
                            "SELECT pr.youtube_video_id, pr.keyword, mt.seq "
                            "FROM production_results pr "
                            "LEFT JOIN manual_topics mt ON mt.seq = CAST("
                            "  REPLACE(pr.topic_id, 'manual_', '') AS INTEGER) "
                            "WHERE pr.youtube_video_id IS NOT NULL "
                            "AND pr.youtube_video_id != ''"
                        ).fetchall()
                    else:
                        rows = conn.execute(
                            "SELECT youtube_video_id, keyword, NULL "
                            "FROM production_results "
                            "WHERE youtube_video_id IS NOT NULL "
                            "AND youtube_video_id != ''"
                        ).fetchall()

                    for vid_id, keyword, seq in rows:
                        seq_str = str(seq) if seq is not None else ""
                        vid_lookup[vid_id] = (seq_str, keyword or "")
                finally:
                    conn.close()
            except Exception as exc:
                logger.warning("[harvest] Could not query production_results: %s", exc)

        data_rows = []
        for v in videos:
            seq, brief = vid_lookup.get(v.video_id, ("", ""))
            data_rows.append([
                v.video_id,
                v.published_at[:10] if v.published_at else "",
                seq,
                brief,
                v.title,
                v.views,
                v.likes,
                round(v.average_view_percentage, 1),
                round(v.average_view_duration, 1),
                v.subscribers_gained,
                v.shares,
                v.category_id,
            ])

        if data_rows:
            ws.append_rows(data_rows, value_input_option="USER_ENTERED")

        logger.info("[harvest] Performance tab synced — %d rows", len(data_rows))

    def _get_channel_name(self) -> str:
        """Return a filename-safe channel name (e.g. 'MoneyHeresy')."""
        try:
            from config.channels import CHANNELS  # lazy
            for ch in CHANNELS:
                if ch.channel_key == self.channel_key:
                    return ch.name.replace(" ", "")
        except Exception:
            pass
        return self.channel_key.replace("_", " ").title().replace(" ", "")

    def _header_style(self):
        from openpyxl.styles import Font, PatternFill
        fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        font = Font(color="FFFFFF", bold=True)
        return fill, font

    def _write_channel_overview_sheet(self, ws, stats: ChannelStats) -> None:
        """Sheet 1: channel stats table + last-7-harvests trend."""
        from openpyxl.styles import Font

        header_fill, header_font = self._header_style()

        # Stats table
        for col, h in enumerate(["Metric", "Value"], 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font

        data = [
            ("Date",                 stats.date),
            ("Subscribers",          stats.subscribers),
            ("Total Views",          stats.total_views),
            ("Total Videos",         stats.total_videos),
            ("Watch Time (hours)",   stats.watch_time_hours),
        ]
        for r, (metric, value) in enumerate(data, 2):
            ws.cell(row=r, column=1, value=metric)
            ws.cell(row=r, column=2, value=value)

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

        # Trend section
        ws.cell(row=9, column=1, value="Recent History (last 7 harvests)").font = Font(bold=True)
        recent = self._load_recent_channel_stats(7)
        if recent:
            trend_cols = ["Date", "Subscribers", "Total Views", "Total Videos"]
            for col, h in enumerate(trend_cols, 1):
                cell = ws.cell(row=10, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
            for r, row in enumerate(recent, 11):
                ws.cell(row=r, column=1, value=row.get("date", ""))
                ws.cell(row=r, column=2, value=row.get("subscribers", ""))
                ws.cell(row=r, column=3, value=row.get("total_views", ""))
                ws.cell(row=r, column=4, value=row.get("total_videos", ""))

    def _write_video_metrics_sheet(self, ws, videos: list[VideoRow]) -> None:
        """
        Sheet 2: all videos sorted by views desc.
        Conditional formatting: views > 1000 green, 500-999 amber, < 100 red.
        Auto-filter on all columns.
        """
        from openpyxl.styles import PatternFill
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.utils import get_column_letter

        header_fill, header_font = self._header_style()

        for col, name in enumerate(VIDEO_CSV_COLUMNS, 1):
            cell = ws.cell(row=1, column=col, value=name)
            cell.fill = header_fill
            cell.font = header_font

        sorted_videos = sorted(videos, key=lambda v: v.views, reverse=True)
        for r, video in enumerate(sorted_videos, 2):
            d = video.to_dict()
            for col, name in enumerate(VIDEO_CSV_COLUMNS, 1):
                ws.cell(row=r, column=col, value=d.get(name, ""))

        # Conditional formatting on the "views" column
        views_col_idx    = VIDEO_CSV_COLUMNS.index("views") + 1
        views_col_letter = get_column_letter(views_col_idx)
        last_row         = max(len(videos) + 1, 2)
        views_range      = f"{views_col_letter}2:{views_col_letter}{last_row}"

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        amber_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        red_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        ws.conditional_formatting.add(
            views_range,
            CellIsRule(operator="greaterThan", formula=["1000"], fill=green_fill),
        )
        ws.conditional_formatting.add(
            views_range,
            CellIsRule(operator="between", formula=["500", "999"], fill=amber_fill),
        )
        ws.conditional_formatting.add(
            views_range,
            CellIsRule(operator="lessThan", formula=["100"], fill=red_fill),
        )

        # Auto-filter covering all data
        ws.auto_filter.ref = ws.dimensions

    def _write_top_performers_sheet(self, ws, videos: list[VideoRow]) -> None:
        """Sheet 3: top-10 tables for views, watch-time %, and like rate."""
        current_row = 1

        def write_section(
            title: str,
            items: list[VideoRow],
            value_col: str,
            value_fn,
        ) -> None:
            nonlocal current_row
            from openpyxl.styles import Font

            header_fill, header_font = self._header_style()
            ws.cell(row=current_row, column=1, value=title).font = Font(bold=True)
            current_row += 1

            for col, h in enumerate(["Rank", "Video ID", "Title", value_col], 1):
                cell = ws.cell(row=current_row, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
            current_row += 1

            for rank, video in enumerate(items, 1):
                ws.cell(row=current_row, column=1, value=rank)
                ws.cell(row=current_row, column=2, value=video.video_id)
                ws.cell(row=current_row, column=3, value=video.title)
                ws.cell(row=current_row, column=4, value=value_fn(video))
                current_row += 1

            current_row += 1  # blank row between sections

        top_views = sorted(videos, key=lambda v: v.views, reverse=True)[:10]
        top_watch = sorted(
            videos, key=lambda v: v.average_view_percentage, reverse=True
        )[:10]
        top_likes = sorted(videos, key=lambda v: v.like_rate, reverse=True)[:10]

        write_section("Top 10 by Views",        top_views, "Views",     lambda v: v.views)
        write_section("Top 10 by Watch Time %", top_watch, "Avg View %", lambda v: v.average_view_percentage)
        write_section("Top 10 by Like Rate",    top_likes, "Like Rate",  lambda v: round(v.like_rate, 4))

    def _load_recent_channel_stats(self, max_count: int) -> list[dict[str, Any]]:
        """Load rows from up to max_count recent channel_stats CSV files."""
        import csv
        import glob as glob_mod

        pattern = str(self.output_dir / "channel_stats_*.csv")
        files   = sorted(glob_mod.glob(pattern), reverse=True)[:max_count]
        rows: list[dict[str, Any]] = []
        for filepath in files:
            try:
                with open(filepath, newline="", encoding="utf-8") as fh:
                    for row in csv.DictReader(fh):
                        rows.append(row)
            except Exception as exc:
                logger.warning("[harvest] Could not read %s: %s", filepath, exc)
        return rows[:max_count]


# ---------------------------------------------------------------------------
# Top-level function — safe to call from scheduler
# ---------------------------------------------------------------------------

def harvest(
    channel: str = "money_debate",
    output_dir: str | Path = DEFAULT_OUTPUT_DIR,
    format: str = "both",
) -> dict[str, Any]:
    """
    Harvest YouTube analytics for the given channel.

    Args:
        channel:    Channel key — must match a .credentials/{channel}_token.json file.
        output_dir: Directory to save CSV and/or Excel output.
        format:     "csv", "excel", or "both".

    Returns:
        HarvestResult.to_dict() on success.
        {"channel_key": channel, "is_valid": False, "error": str} on unexpected failure.
        Never raises — safe to call from the APScheduler job.
    """
    try:
        harvester = AnalyticsHarvester(
            channel_key=channel,
            output_dir=Path(output_dir),
        )
        result = harvester.harvest(format=format)
        print(result.summary())
        return result.to_dict()
    except Exception as exc:
        logger.error("[harvest] Unexpected failure for channel '%s': %s", channel, exc)
        return {"channel_key": channel, "is_valid": False, "error": str(exc)}


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Harvest YouTube analytics and save to CSV + Excel."
    )
    p.add_argument("--channel", default="money_debate",
                   help="Channel key (default: money_debate)")
    p.add_argument("--output",  default=str(DEFAULT_OUTPUT_DIR),
                   help="Output directory (default: data/analytics/)")
    p.add_argument("--format",  choices=["csv", "excel", "both"], default="both",
                   help="Output format (default: both)")
    return p


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    args = _build_arg_parser().parse_args()
    result = harvest(channel=args.channel, output_dir=args.output, format=args.format)
    if not result.get("is_valid", True):
        sys.exit(1)
